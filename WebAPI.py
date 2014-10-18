########################
# WebAPI Python Script for AtomPy 2.0
# Created by Josiah Lucas Boswell (www.josiahboswell.com)
# Is called by a PHP script with parameters for Z and N,
# a paramater for the sheet number, and a backup argument.
# It then returns the HTML code from a webpage displayment
# of the requested file.
########################

#Import necessary plugins
import openpyxl, os, sys, time
import sys, json, locale
locale.setlocale(locale.LC_ALL, '')

#Get Z, N, and SheetNum values from the PHP post
Z = int(sys.argv[1])
N = int(sys.argv[2])
SheetNum = int(sys.argv[3])
BackupArg = str(sys.argv[4])

#Build the filename string
filename = ''
if Z < 10:
	filename = filename + '0' + str(Z)
else:
	filename = filename + str(Z)
filename = filename + '_'
if N < 10:
	filename = filename + '0' + str(N)
else:
	filename = filename + str(N)
filename = filename + '.xlsx'

#Now open the file
wb = None
if BackupArg == str(-1):
	wb = openpyxl.load_workbook('Database//' + filename)
else:
	wb = openpyxl.load_workbook('Backups//' + BackupArg + '//' + filename)

#Get the sheet
ws = wb.get_sheet_by_name(wb.get_sheet_names()[SheetNum])

#Get the column widths
column_widths = []
for i in range(len(ws.columns)):
	if ws.column_dimensions[openpyxl.cell.get_column_letter(i+1)].width != None:
		column_widths.append(ws.column_dimensions[openpyxl.cell.get_column_letter(i+1)].width)

#Get the headerRow
headerRow = -1
for i in range(len(ws.rows)):
	if ws.cell(row = i+1, column = 1).value == 'Z':
		headerRow = i-1
		break

#Create our webpage string which will be output to the PHP script
#when we are all done
webpage = ''

#Display what version of the database we are on
webpage += '<br>Current Displaying: '
if BackupArg == '-1':
	webpage += 'Most Recent Version'
else:
	webpage += str(BackupArg)

#Add backup revisions selection table
avaliableBackups = []
backupFolders = os.listdir('Backups//')
backupFolders = [x for x in backupFolders if 'old' not in x]
for i in range(len(backupFolders)):
	if filename in os.listdir('Backups//' + backupFolders[i]):
		avaliableBackups.append(backupFolders[i])
if len(avaliableBackups) == 0:
	webpage += '<br>No backups for this data sheet.<br>'
else:
	webpage += '<br>Backups avaliable:<br>'
	webpage += '<form action="viewFile.php" method="post">'
	webpage += '<input type="hidden" name="Z" value=' + str(Z)
	webpage += '><input type="hidden" name="N" value=' + str(N)
	webpage += '><input type="hidden" name="SheetNum" value=' + str(SheetNum)
	webpage += '><select name="BackupArg">'

	#Show option for the current version of the file
	webpage += '<option name="BackupArg" value="-1">Current Version</option>'

	#Show options for the last versions of the file
	for i in range(len(avaliableBackups)):
		webpage += '<option name="BackupArg" value="'
		webpage += avaliableBackups[i] + '">'
		webpage += avaliableBackups[i] + '</option>'
	webpage += '</select><br><input type="submit" value="Submit"></form>'

#Add the sheet selection table
webpage += '<table style="width:300px"><tr>'
for i in range(len(wb.get_sheet_names())):
	webpage += '<td><form action="viewFile.php" method="post">'
	webpage += '<input type="hidden" name="Z" value=' + str(Z)
	webpage += '><input type="hidden" name="N" value=' + str(N)
	webpage += '><input type="hidden" name="SheetNum" value=' + str(i)
	webpage += '><input type="hidden" name="BackupArg" value="-1">'
	webpage += '<input type="submit" value="' + wb.get_sheet_by_name(wb.get_sheet_names()[i]).title
	webpage += '"></form></td>'
webpage += '</tr></table><br><br>'

#Debug print
webpage += "Retrieved " + str(len(ws.columns)*len(ws.rows)) + " cells...<br>"

#Display cell data (header and everything)
counter = 0;
webpage += '<table>'
for i in range(len(column_widths)):
	webpage += "<col width='" + str(column_widths[i]*10) + "'>"

for i in range(len(ws.rows)):
	webpage += "<tr>"
	for j in range(len(ws.columns)):
		#First, get out current cell (cCell)
		cCell = ws.cell(row = i+1, column = j+1)
		
		#Begin column with text alignment and spanning
		webpage += "<td id='"
		if cCell.style.number_format == 'General':
			webpage += "left'"
		else:
			webpage += "right'"
		if i < headerRow-1:
			webpage += " colspan='"
			webpage += str(len(ws.columns))
			webpage += "'"
		webpage += ">"
		
		#Begin creating the css style for this cell
		if cCell.value == '' or cCell.value == None:
			#No styling needed, print out nothing but the row end
			webpage += "</td>"
		else:
			webpage += "<style>custom"
			webpage += str(counter)
			
			#Font family and font size
			webpage += "{font-family: '" + cCell.style.font.name
			webpage += "'; font-size:" + str((cCell.style.font.sz/10.0)*100) + "%;"
			
			#Bold
			if cCell.style.font.b == True:
				webpage += "font-weight: bold;"
			else:
				webpage += "font-weight: normal;"
			
			#Italic
			if cCell.style.font.i == True:
				webpage += "font-style: italic;"
			else:
				webpage += "font-style: normal;"
			
			#Finish off the css style for this cell
			webpage += "}</style><custom" + str(counter) + ">"
			
			#Add decimals, commas to numbers
			if not (type(cCell.value) is unicode):
				if not (cCell.style.number_format == 'Scientific'):
					cCell.value = locale.format("%.2f", cCell.value, grouping= True)
				else:
					cCell.value = locale.format("%.4E", cCell.value, grouping= True)
			
			#Hyperlink handling
			if type(cCell.value) is unicode:
				if 'HYPERLINK' in cCell.value:
					temp = cCell.value.split('"')
					cCell.value = '<a href="' + temp[1] + '">' + temp[3] + '</a>';
			
			#Print out the cell value
			if type(cCell.value) is unicode:
				cCell.value = cCell.value.replace(u'\xa0','')
			webpage += str(cCell.value)
			
			#End the column
			webpage += "</custom" + str(counter) + "></td>";
			counter = counter + 1;
		
		#Skip to next row if this cell was a header row
		if i < headerRow:
			break
			
	webpage += "</tr>"

webpage += "</table>"

#Return the webpage to the PHP script
print json.dumps(webpage)
