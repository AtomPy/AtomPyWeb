########################
# ViewFile Python Script for AtomPy 2.1
#
# Created by Josiah Lucas Boswell (www.josiahboswell.com)
#
# Called from a PHP script with two arguments, filename and sheet number.
#
########################

#Import necessary plugins
import openpyxl, os, sys, time, locale, json
locale.setlocale(locale.LC_ALL, '')

#Get filename and sheet number from the PHP script
Z = int(sys.argv[1])
N = int(sys.argv[2])
SheetNum = int(sys.argv[3])
BackupArg = int(sys.argv[4])
BackupsString = str(sys.argv[5])
Backups = BackupsString.split(',')

#Build filename
filename = ''
if(Z < 10):
	filename += '0' + str(Z)
else:
	filename += str(Z)
filename +=  '_'
if(N < 10):
	filename += '0' +str(N)
else:
	filename += str(N)
filename += '.xlsx'

#Now open the file and grab the sheet
wb = openpyxl.load_workbook('Database//' + filename)
ws = wb.get_sheet_by_name(wb.get_sheet_names()[SheetNum])

#Get the column widths
column_widths = []
for i in range(len(ws.columns)):
	if ws.column_dimensions[openpyxl.cell.get_column_letter(i+1)].width != None:
		column_widths.append(ws.column_dimensions[openpyxl.cell.get_column_letter(i+1)].width)

#Get the headerRow
headerRow = -1
for i in range(len(ws.rows)):
	if ws.cell(row = i+1, column = 1).value == 'Z':
		headerRow = i-1
		break

#Create our webpage string which will be output to the PHP script
#when we are all done
webpage = '<br>'

#Display what version of the file we are on and give a select for all available backups of the file
webpage += '<form action="viewFile.php" method="post">File Version: <select name="BackupArg">'

#Display the current version as first in the list and then show the remaining ones
if(BackupArg == -1):
	webpage += '<option value="-1">Most Recent Version</option>'
	for i in range(len(Backups)):
		webpage += '<option value="' + str(i) + '">' + str(Backups[i]) + '</option>'
else:
	webpage += '<option value="' + str(BackupArg) + '">' + str(Backups[BackupArg]) + '</option>'
	webpage += '<option value="-1">Most Recent Version</option>'
	for i in range(len(Backups)):
		if i != BackupArg:
			webpage += '<option value="' + str(i) + '">' + str(Backups[i]) + '</option>'

#Finish off the select form
webpage += '</select><input type="hidden" name="Z" value=' + str(Z)
webpage += '><input type="hidden" name="N" value=' + str(N)
webpage += '><input type="hidden" name="SheetNum" value=' + str(SheetNum)
webpage += '><input type="submit" value="Load Backup"></form>'

#Selection of sheets
webpage += '<form action="viewFile.php" method="post">Select Sheet: <select name="SheetNum">'
for i in range(len(wb.get_sheet_names())):
	webpage += '<option value="' + str(i) + '">' + wb.get_sheet_names()[i] + '</option>'
webpage += '</select><input type="hidden" name="Z" value=' + str(Z)
webpage += '><input type="hidden" name="N" value=' + str(N)
webpage += '><input type="hidden" name="BackupArg" value=' + str(BackupArg)
webpage += '><input type="submit" value="Load Sheet"></form>'

#Debug print
webpage += "Retrieved " + str(len(ws.columns)*len(ws.rows)) + " cells...<br>"

#Display cell data (header and everything)
webpage += '<table>'
for i in range(len(column_widths)):
	webpage += "<col width='" + str(column_widths[i]*10) + "'>"

#Go through all of the cells and do the CSS stylings
for i in range(len(ws.rows)):
	webpage += "<tr>"
	for j in range(len(ws.columns)):
		
		#First, get out current cell (cCell)
		cCell = ws.cell(row = i+1, column = j+1)
		
		#Begin column with text alignment and spanning
		webpage += "<td "
		if i < headerRow-1:
			webpage += " colspan='" + str(len(ws.columns)) + "'"
		webpage += ">"
		
		#Begin creating the css style for this cell
		if cCell.value == '' or cCell.value == None:
			#No styling needed, print out nothing but the row end
			webpage += "</td>"
		else:
			#Add decimals, commas to numbers
			if not (type(cCell.value) is unicode):
				if not (cCell.style.number_format == 'Scientific'):
					cCell.value = locale.format("%.2f", cCell.value, grouping=True)
				else:
					cCell.value = locale.format("%.4E", cCell.value, grouping=True)
			
			#Hyperlink handling
			if type(cCell.value) is unicode:
				if 'HYPERLINK' in cCell.value:
					temp = cCell.value.split('"')
					cCell.value = '<a href="' + temp[1] + '">' + temp[3] + '</a>';
			
			#Print out the cell value
			if type(cCell.value) is unicode:
				cCell.value = cCell.value.replace(u'\xa0','')
			webpage += str(cCell.value)
			
			#End the column
			webpage += "</td>";
		
		#Skip to next row if this cell was a header row
		if i < headerRow:
			break
			
	webpage += "</tr>"

webpage += "</table>"

#Return the webpage to the PHP script
print webpage
